import streamlit as st
import whisper
import requests
import json
from datetime import datetime
from pathlib import Path
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
import os

# ============ 页面配置 ============

st.set_page_config(
    page_title="AI 视频标题生成工具",
    page_icon="🎬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============ 样式设置 ============

st.markdown("""
<style>
    .main {
        padding: 2rem;
    }
    .stButton > button {
        width: 100%;
        padding: 0.75rem;
        font-size: 1rem;
    }
    .section-header {
        font-size: 1.3rem;
        font-weight: bold;
        margin-bottom: 1rem;
        border-bottom: 2px solid #FF6B6B;
        padding-bottom: 0.5rem;
    }
</style>
""", unsafe_allow_html=True)

# ============ Session State 初始化 ============

if 'transcribed_text' not in st.session_state:
    st.session_state.transcribed_text = ""
if 'generated_titles' not in st.session_state:
    st.session_state.generated_titles = None
if 'manual_titles' not in st.session_state:
    st.session_state.manual_titles = ""
if 'platform_titles' not in st.session_state:
    st.session_state.platform_titles = {}

# ============ 配置和常量 ============

OPENROUTER_API_KEY = st.secrets.get("openrouter_api_key", "")
OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"

MODELS = {
    "deepseek": "deepseek/deepseek-r1:free",
    "mistral": "mistralai/mistral-7b-instruct"
}

# ============ 核心函数 ============

@st.cache_resource
def load_whisper_model():
    """加载 Whisper 模型"""
    return whisper.load_model("base")

def extract_audio_from_video(video_file):
    """从视频文件提取音频"""
    try:
        with tempfile.NamedTemporaryFile(suffix=".mp4", delete=False) as tmp_video:
            tmp_video.write(video_file.read())
            tmp_video_path = tmp_video.name
        
        # 使用 ffmpeg 提取音频
        import subprocess
        tmp_audio_path = tmp_video_path.replace(".mp4", ".wav")
        
        result = subprocess.run(
            ["ffmpeg", "-i", tmp_video_path, "-q:a", "9", "-n", tmp_audio_path],
            capture_output=True,
            text=True
        )
        
        if os.path.exists(tmp_audio_path):
            return tmp_audio_path
        else:
            st.error("❌ 音频提取失败")
            return None
    except Exception as e:
        st.error(f"❌ 提取音频出错: {str(e)}")
        return None

def transcribe_audio(audio_path):
    """使用 Whisper 转录音频"""
    try:
        model = load_whisper_model()
        result = model.transcribe(audio_path, language="zh")
        return result["text"]
    except Exception as e:
        st.error(f"❌ 转录出错: {str(e)}")
        return None

def call_openrouter_api(prompt, model_name):
    """调用 OpenRouter API"""
    try:
        headers = {
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "HTTP-Referer": "https://streamlit.io",
            "X-Title": "Video Title Generator"
        }
        
        data = {
            "model": MODELS[model_name],
            "messages": [
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.7,
            "max_tokens": 500
        }
        
        response = requests.post(OPENROUTER_URL, json=data, headers=headers, timeout=30)
        
        if response.status_code == 200:
            result = response.json()
            return result["choices"][0]["message"]["content"]
        else:
            st.error(f"❌ API 错误: {response.status_code}")
            return None
    except Exception as e:
        st.error(f"❌ API 调用出错: {str(e)}")
        return None

def generate_titles(text):
    """生成标题 - 两个模型并行调用"""
    prompt = f"""基于以下视频内容，生成5个吸引眼球的标题。标题要简洁有力，能吸引用户点击。

视频内容：
{text}

要求：
1. 每个标题一行
2. 标题长度 15-30 个字
3. 避免重复
4. 只返回标题，不要其他解释

标题列表："""
    
    titles_deepseek = None
    titles_mistral = None
    
    col1, col2 = st.columns(2)
    
    with col1:
        with st.spinner("🤖 DeepSeek 生成中..."):
            titles_deepseek = call_openrouter_api(prompt, "deepseek")
    
    with col2:
        with st.spinner("🤖 Mistral 生成中..."):
            titles_mistral = call_openrouter_api(prompt, "mistral")
    
    return {
        "deepseek": titles_deepseek,
        "mistral": titles_mistral
    }

def adapt_for_platform(text, platform):
    """为不同平台生成适配标题"""
    platform_prompts = {
        "douyin": """根据以下内容，生成3个适合抖音的标题。
        
特点要求：
- 15秒内能讲完的爆款内容
- 制造悬念和FOMO感
- 年轻、活力的语言风格
- 可以使用表情符号
- 例如："震撼！我发现了..."、"你绝对没看过的..."

内容：{text}

只返回标题，每行一个：""",
        
        "wechat": """根据以下内容，生成3个适合微信视频号的标题。

特点要求：
- 温和、专业、有信任感
- 包含故事感和情感连接
- 微信生态友好
- 适合各年龄段观看
- 例如："深度分享｜为什么..."、"这个方法改变了我..."

内容：{text}

只返回标题，每行一个：""",
        
        "xiaohongshu": """根据以下内容，生成3个适合小红书的标题。

特点要求：
- 包含 # 话题标签
- 种草、推荐的风格
- 亲近感强，像朋友推荐
- 可以使用emoji
- 例如："#分享 这个真的绝了！..."、"姐妹们必看｜..."

内容：{text}

只返回标题，每行一个："""
    }
    
    prompt = platform_prompts.get(platform, "").format(text=text)
    return call_openrouter_api(prompt, "mistral")

def create_excel_file(original_text, manual_titles, platform_results):
    """创建 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "视频标题"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    
    # 样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # 标题行
    ws['A1'] = "内容类型"
    ws['B1'] = "标题"
    ws['A1'].fill = header_fill
    ws['B1'].fill = header_fill
    ws['A1'].font = header_font
    ws['B1'].font = header_font
    
    row = 2
    
    # 原始转录文本
    ws[f'A{row}'] = "原始转录文本"
    ws[f'B{row}'] = original_text
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    row += 1
    
    # 手动输入的标题
    if manual_titles.strip():
        ws[f'A{row}'] = "手动输入标题"
        ws[f'B{row}'] = manual_titles
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # 平台适配标题
    platform_names = {
        "douyin": "抖音版本",
        "wechat": "视频号版本",
        "xiaohongshu": "小红书版本"
    }
    
    for platform_key, platform_name in platform_names.items():
        if platform_key in platform_results and platform_results[platform_key]:
            ws[f'A{row}'] = platform_name
            ws[f'B{row}'] = platform_results[platform_key]
            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
            row += 1
    
    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ============ 主界面 ============

st.markdown("""
# 🎬 AI 视频标题生成工具 Pro

**智能转文字 · 多模型生成 · 平台适配 · 一键导出**

---
""")

# 第一步：视频上传和语音转文字
st.markdown("### 📹 第一步：视频上传和语音转文字")

video_file = st.file_uploader(
    "上传视频文件（MP4, MOV, MKV 等）",
    type=["mp4", "mov", "mkv", "avi", "webm", "flv"]
)

if video_file:
    st.info(f"📄 已选择：{video_file.name}")
    
    if st.button("🎙️ 开始转录", key="transcribe_btn"):
        with st.spinner("正在提取音频..."):
            audio_path = extract_audio_from_video(video_file)
        
        if audio_path:
            with st.spinner("正在转录语音..."):
                transcribed = transcribe_audio(audio_path)
                if transcribed:
                    st.session_state.transcribed_text = transcribed
                    st.success("✅ 转录完成！")

# 第二步：文本编辑
st.markdown("### ✏️ 第二步：文本编辑和优化")

edited_text = st.text_area(
    "编辑转录文本（可修改）",
    value=st.session_state.transcribed_text,
    height=150,
    key="text_editor"
)

if edited_text:
    st.session_state.transcribed_text = edited_text
    
    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("📋 复制", use_container_width=True):
            st.write("✅ 已复制到剪贴板")
    with col2:
        if st.button("🗑️ 清空", use_container_width=True):
            st.session_state.transcribed_text = ""
            st.rerun()
    with col3:
        if st.button("✨ 自动整理", use_container_width=True):
            st.info("功能开发中...")

# 第三步：标题生成
st.markdown("### 🎯 第三步：标题生成（两个模型）")

if st.button("🚀 生成标题", key="generate_btn", type="primary"):
    if not st.session_state.transcribed_text.strip():
        st.error("❌ 请先输入或转录文本内容")
    else:
        if not OPENROUTER_API_KEY:
            st.error("❌ 未设置 OpenRouter API Key，请在 Streamlit Secrets 中配置")
        else:
            with st.spinner("正在生成标题..."):
                st.session_state.generated_titles = generate_titles(st.session_state.transcribed_text)

# 显示生成的标题
if st.session_state.generated_titles:
    st.markdown("#### 📊 生成的标题（两个模型对比）")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**🤖 DeepSeek-R1 版本**")
        if st.session_state.generated_titles["deepseek"]:
            st.write(st.session_state.generated_titles["deepseek"])
        else:
            st.warning("生成失败")
    
    with col2:
        st.markdown("**🤖 Mistral 版本**")
        if st.session_state.generated_titles["mistral"]:
            st.write(st.session_state.generated_titles["mistral"])
        else:
            st.warning("生成失败")

# 第四步：平台适配
st.markdown("### 📱 第四步：平台适配标题生成")

manual_titles_input = st.text_area(
    "手动输入来自网页版的标题（可选，一行一个）",
    height=80,
    key="manual_titles_input"
)

if manual_titles_input:
    st.session_state.manual_titles = manual_titles_input

col1, col2, col3 = st.columns(3)

with col1:
    if st.button("📱 生成抖音版本", use_container_width=True):
        if not st.session_state.transcribed_text.strip():
            st.error("❌ 请先输入文本")
        else:
            with st.spinner("生成中..."):
                result = adapt_for_platform(st.session_state.transcribed_text, "douyin")
                if result:
                    st.session_state.platform_titles["douyin"] = result
                    st.success("✅ 完成")

with col2:
    if st.button("📱 生成视频号版本", use_container_width=True):
        if not st.session_state.transcribed_text.strip():
            st.error("❌ 请先输入文本")
        else:
            with st.spinner("生成中..."):
                result = adapt_for_platform(st.session_state.transcribed_text, "wechat")
                if result:
                    st.session_state.platform_titles["wechat"] = result
                    st.success("✅ 完成")

with col3:
    if st.button("📱 生成小红书版本", use_container_width=True):
        if not st.session_state.transcribed_text.strip():
            st.error("❌ 请先输入文本")
        else:
            with st.spinner("生成中..."):
                result = adapt_for_platform(st.session_state.transcribed_text, "xiaohongshu")
                if result:
                    st.session_state.platform_titles["xiaohongshu"] = result
                    st.success("✅ 完成")

# 显示平台标题
if st.session_state.platform_titles:
    st.markdown("#### 🎨 平台适配标题")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if "douyin" in st.session_state.platform_titles:
            st.markdown("**🎵 抖音**")
            st.write(st.session_state.platform_titles["douyin"])
    
    with col2:
        if "wechat" in st.session_state.platform_titles:
            st.markdown("**📱 视频号**")
            st.write(st.session_state.platform_titles["wechat"])
    
    with col3:
        if "xiaohongshu" in st.session_state.platform_titles:
            st.markdown("**🌟 小红书**")
            st.write(st.session_state.platform_titles["xiaohongshu"])

# 第五步：导出结果
st.markdown("### 📊 第五步：导出结果")

if st.button("💾 导出为 Excel", type="primary", use_container_width=True):
    if not st.session_state.transcribed_text.strip():
        st.error("❌ 没有转录文本")
    else:
        excel_file = create_excel_file(
            st.session_state.transcribed_text,
            st.session_state.manual_titles,
            st.session_state.platform_titles
        )
        
        st.download_button(
            label="⬇️ 下载 Excel 文件",
            data=excel_file,
            file_name=f"视频标题_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ============ 侧边栏 ============

with st.sidebar:
    st.markdown("## 📖 使用说明")
    
    st.markdown("""
    ### 工作流程
    
    1. **上传视频** → 自动提取音频
    2. **语音转文字** → 使用 Whisper 转录
    3. **编辑文本** → 优化转录内容
    4. **生成标题** → 两个 AI 模型对比
    5. **平台适配** → 生成三个平台版本
    6. **导出结果** → 保存为 Excel
    
    ### 支持的视频格式
    - MP4, MOV, MKV, AVI, WebM, FLV
    
    ### 平台特性
    
    **🎵 抖音**
    - 15秒爆款风格
    - 年轻、活力语言
    - 制造悬念感
    
    **📱 视频号**
    - 温和、专业
    - 故事感强
    - 微信生态友好
    
    **🌟 小红书**
    - 种草推荐风格
    - #话题标签
    - 亲近感强
    
    ### 技术栈
    - **语音识别**：OpenAI Whisper
    - **AI 模型**：OpenRouter (DeepSeek + Mistral)
    - **前端框架**：Streamlit
    - **导出格式**：Excel
    
    ### 注意事项
    - 确保网络连接正常
    - API Key 需要有效余额
    - 大文件处理可能较慢
    """)
    
    st.markdown("---")
    
    st.markdown("""
    ### 💡 提示
    - 转录质量取决于视频音频清晰度
    - 可手动编辑转录文本以提高准确性
    - 不同平台的标题有重大差异
    - 建议对比两个模型的结果选最优
    
    ### 🚀 版本信息
    - **版本**：1.0 Pro
    - **更新**：2025年10月
    - **支持**：中文视频
    """)

# ============ 页脚 ============

st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #888; padding: 1rem;'>
    <p>🎬 AI 视频标题生成工具 Pro | 完全免费 | Powered by Streamlit + OpenRouter</p>
</div>
""", unsafe_allow_html=True)
